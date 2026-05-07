from flask import Flask, render_template, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from datetime import datetime, timedelta
import json
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
import random

app = Flask(__name__)

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///pickleball.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'pickleball-dsp-secret-key-2026'
db = SQLAlchemy(app)
migrate = Migrate(app, db)

import unicodedata

def clean_player_name(value):
    if not value:
        return ''

    text = str(value).strip()

    if '!' in text:
        text = text.split('!')[0].strip()

    return text

def normalize_name(value):
    if not value:
        return ''
    
    text = str(value).strip().upper()
    
    # 🔥 Supprime les accents correctement
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(c for c in text if not unicodedata.combining(c))
    
    # Nettoyage des espaces
    return ' '.join(text.split())

# ==================== MODELS ====================

class Player(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String(100), nullable=False)
    last_name = db.Column(db.String(100), nullable=False)
    full_name = db.Column(db.String(200), nullable=False)
    gender = db.Column(db.String(1), default='M')
    level = db.Column(db.Float, default=3.5)
    email = db.Column(db.String(200))
    status = db.Column(db.String(50), default='Actif')
    selected = db.Column(db.Boolean, default=False)
    drill = db.Column(db.Boolean, default=False)
    
    def to_dict(self):
        return {
            'id': self.id,
            'firstName': self.first_name,
            'lastName': self.last_name,
            'fullName': self.full_name,
            'gender': self.gender,
            'level': self.level,
            'email': self.email,
            'status': self.status,
            'selected': self.selected,
            'drill': self.drill
        }

class Event(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    client = db.Column(db.String(200))
    day = db.Column(db.String(50))
    start_time = db.Column(db.String(10))
    end_time = db.Column(db.String(10))
    drill_minutes = db.Column(db.Integer, default=0)
    period_duration = db.Column(db.Integer, default=20)
    cost = db.Column(db.Float, default=0.0)
    
    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'client': self.client,
            'day': self.day,
            'startTime': self.start_time,
            'endTime': self.end_time,
            'drillMinutes': self.drill_minutes,
            'periodDuration': self.period_duration,
            'cost': self.cost
        }

class Schedule(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    event_id = db.Column(db.Integer, db.ForeignKey('event.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    selected_players = db.Column(db.Text)  # JSON array of player IDs
    drill_players = db.Column(db.Text)  # JSON array of player IDs
    schedule_data = db.Column(db.Text)  # JSON of complete schedule
    
    event = db.relationship('Event', backref='schedules')

class Setting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.String(200))
    
    @staticmethod
    def get(key, default=None):
        setting = Setting.query.filter_by(key=key).first()
        return setting.value if setting else default
    
    @staticmethod
    def set(key, value):
        setting = Setting.query.filter_by(key=key).first()
        if setting:
            setting.value = str(value)
        else:
            setting = Setting(key=key, value=str(value))
            db.session.add(setting)
        db.session.commit()

# ==================== ROUTES ====================
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/players', methods=['GET'])
def get_players():
    players = Player.query.all()
    return jsonify([p.to_dict() for p in players])

@app.route('/api/events', methods=['GET'])
def get_events():
    events = Event.query.all()
    return jsonify([e.to_dict() for e in events])

@app.route('/api/settings', methods=['GET'])
def get_settings():
    return jsonify({
        'maxTeammates': int(Setting.get('maxTeammates', 1)),
        'maxOpponents': int(Setting.get('maxOpponents', 2)),
        'maxTeamLevelDiff': float(Setting.get('maxTeamLevelDiff', 0.4)),
        'maxMatchLevelDiff': float(Setting.get('maxMatchLevelDiff', 0.49))
    })

@app.route('/api/settings', methods=['POST'])
def update_settings():
    data = request.json
    for key, value in data.items():
        Setting.set(key, value)
    return jsonify({'success': True})

@app.route('/api/reset', methods=['POST'])
def reset_data():
    try:
        # Supprimer toutes les données
        Schedule.query.delete()
        Event.query.delete()
        Player.query.delete()
        Setting.query.delete()
        
        db.session.commit()

        # Recréer les paramètres par défaut
        Setting.set('maxTeammates', 1)
        Setting.set('maxOpponents', 2)
        Setting.set('maxTeamLevelDiff', 0.4)
        Setting.set('maxMatchLevelDiff', 0.49)

        return jsonify({'success': True, 'message': 'Base complètement remise à zéro'})
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/upload', methods=['POST'])

def upload_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file'}), 400

    file = request.files['file']
    wb = openpyxl.load_workbook(file, data_only=True)

    result = {'players': 0, 'events': 0, 'selected': 0, 'drill': 0}

    def normalize_name(value):
        if not value:
            return ''
        return ' '.join(str(value).strip().upper().split())

    # Import players from "Noms", "Sélection joueurs" or "Membres"
    sheet_name = None
    if 'Noms' in wb.sheetnames:
        sheet_name = 'Noms'
    elif 'Sélection joueurs' in wb.sheetnames:
        sheet_name = 'Sélection joueurs'
    elif 'Membres' in wb.sheetnames:
        sheet_name = 'Membres'

    # Lire la feuille Membres pour récupérer les informations complètes (genre, niveau, etc.)
    members_map = {}

    if 'Membres' in wb.sheetnames:
        ws_members = wb['Membres']
        member_headers = [str(cell.value).strip() if cell.value else '' for cell in ws_members[1]]

        for row in ws_members.iter_rows(min_row=2, values_only=True):
            full_name = ''
            first_name = ''
            last_name = ''
            gender = 'M'
            level = 3.5
            email = ''
            status = 'Actif'

            if '(F) Nom complet' in member_headers:
                idx = member_headers.index('(F) Nom complet')
                full_name = str(row[idx]).strip() if len(row) > idx and row[idx] else ''
            elif 'Nom complet' in member_headers:
                idx = member_headers.index('Nom complet')
                full_name = str(row[idx]).strip() if len(row) > idx and row[idx] else ''

            if 'Prénom' in member_headers:
                idx = member_headers.index('Prénom')
                first_name = str(row[idx]).strip() if len(row) > idx and row[idx] else ''

            if 'Nom' in member_headers:
                idx = member_headers.index('Nom')
                last_name = str(row[idx]).strip() if len(row) > idx and row[idx] else ''

            if 'Genre' in member_headers:
                idx = member_headers.index('Genre')
                gender = str(row[idx]).strip().upper() if len(row) > idx and row[idx] else 'M'

            if 'Niveau' in member_headers:
                idx = member_headers.index('Niveau')
                try:
                    level = float(row[idx]) if len(row) > idx and row[idx] not in (None, '') else 3.5
                except:
                    level = 3.5

            if 'Courriel' in member_headers:
                idx = member_headers.index('Courriel')
                email = str(row[idx]).strip() if len(row) > idx and row[idx] else ''

            if 'Statut' in member_headers:
                idx = member_headers.index('Statut')
                status = str(row[idx]).strip() if len(row) > idx and row[idx] else 'Actif'

            if full_name:
                members_map[normalize_name(full_name)] = {
                    'first_name': first_name,
                    'last_name': last_name,
                    'gender': gender,
                    'level': level,
                    'email': email,
                    'status': status
                }

    # Si le fichier uploadé ne contient pas Membres,
    # récupérer les genres déjà connus dans la base
    if not members_map:
        for player in Player.query.all():
            members_map[normalize_name(player.full_name)] = {
                'first_name': player.first_name,
                'last_name': player.last_name,
                'gender': player.gender,
                'level': player.level,
                'email': player.email,
                'status': player.status
            }

    if sheet_name:
        ws = wb[sheet_name]
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]

        Player.query.delete()

        def is_filled(value):
            return value is not None and str(value).strip() != ''

        for row in ws.iter_rows(min_row=2, values_only=True):
            first_name = ''
            last_name = ''
            full_name = ''

            if 'Prénom' in headers:
                idx = headers.index('Prénom')
                first_name = str(row[idx]).strip() if len(row) > idx and row[idx] else ''

            if 'Nom' in headers:
                idx = headers.index('Nom')
                last_name = str(row[idx]).strip() if len(row) > idx and row[idx] else ''

            if 'Nom complet' in headers:
                idx = headers.index('Nom complet')
                full_name = str(row[idx]).strip() if len(row) > idx and row[idx] else ''
            elif len(row) > 0 and row[0]:
                full_name = str(row[0]).strip()

            if not full_name and (first_name or last_name):
                full_name = f"{first_name} {last_name}".strip()

            if not full_name:
                continue

            # récupération depuis Membres
            full_name = clean_player_name(full_name)
            member_info = members_map.get(normalize_name(full_name), {})

            if not member_info:
                print("AUCUN MATCH MEMBRES POUR:", full_name)

            first_name = member_info.get('first_name', first_name)
            last_name = member_info.get('last_name', last_name)
            gender = member_info.get('gender', 'M')
            level = member_info.get('level', 3.5)
            email = member_info.get('email', '')
            status = member_info.get('status', 'Actif')

            selected = False
            if 'Sélectionner (x)' in headers:
                idx = headers.index('Sélectionner (x)')
                selected = len(row) > idx and is_filled(row[idx])

            drill = False
            if 'Drill (x)' in headers:
                idx = headers.index('Drill (x)')
                drill = len(row) > idx and is_filled(row[idx])

            player = Player(
                first_name=first_name,
                last_name=last_name,
                full_name=full_name,
                gender=gender,
                level=level,
                email=email,
                status=status,
                selected=selected,
                drill=drill
            )

            db.session.add(player)
            result['players'] += 1

            if selected:
                result['selected'] += 1
            if drill:
                result['drill'] += 1

        db.session.commit()

    # Import events
    if 'Événements' in wb.sheetnames:
        ws = wb['Événements']
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]

        Event.query.delete()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            name = str(row[0]).strip() if row[0] else ''

            client = ''
            if 'Client' in headers:
                idx = headers.index('Client')
                client = str(row[idx]).strip() if len(row) > idx and row[idx] else ''

            day = ''
            if 'Journée' in headers:
                idx = headers.index('Journée')
                day = str(row[idx]).strip() if len(row) > idx and row[idx] else ''

            start_time = '12:00'
            if 'Heure début' in headers:
                idx = headers.index('Heure début')
                if len(row) > idx and row[idx]:
                    val = row[idx]
                    if isinstance(val, datetime):
                        start_time = val.strftime('%H:%M')
                    else:
                        start_time = str(val)

            end_time = '15:00'
            if 'Heure fin' in headers:
                idx = headers.index('Heure fin')
                if len(row) > idx and row[idx]:
                    val = row[idx]
                    if isinstance(val, datetime):
                        end_time = val.strftime('%H:%M')
                    else:
                        end_time = str(val)

            drill_minutes = 0
            if 'Drill en minutes' in headers:
                idx = headers.index('Drill en minutes')
                try:
                    drill_minutes = int(row[idx]) if len(row) > idx and row[idx] else 0
                except:
                    drill_minutes = 0

            period_duration = 20
            if "Durée d'une partie" in headers:
                idx = headers.index("Durée d'une partie")
                try:
                    period_duration = int(row[idx]) if len(row) > idx and row[idx] else 20
                except:
                    period_duration = 20

            cost = 0.0
            if 'Coût pour une cédule' in headers:
                idx = headers.index('Coût pour une cédule')
                try:
                    cost = float(row[idx]) if len(row) > idx and row[idx] else 0.0
                except:
                    cost = 0.0

            event = Event(
                name=name,
                client=client,
                day=day,
                start_time=start_time,
                end_time=end_time,
                drill_minutes=drill_minutes,
                period_duration=period_duration,
                cost=cost
            )

            db.session.add(event)
            result['events'] += 1

        db.session.commit()

    return jsonify(result)

def calculate_schedule_score(schedule):
    teammate_counts = {}
    opponent_counts = {}
    penalties = 0

    for period in schedule.get('periods', []):
        for court in period.get('courts', []):
            a1 = court['sideA']['player1']['fullName']
            a2 = court['sideA']['player2']['fullName']
            b1 = court['sideB']['player1']['fullName']
            b2 = court['sideB']['player2']['fullName']

            for pair in [(a1, a2), (a2, a1), (b1, b2), (b2, b1)]:
                teammate_counts[pair] = teammate_counts.get(pair, 0) + 1

            for pair in [
                (a1, b1), (a1, b2),
                (a2, b1), (a2, b2),
                (b1, a1), (b1, a2),
                (b2, a1), (b2, a2)
            ]:
                opponent_counts[pair] = opponent_counts.get(pair, 0) + 1

    for count in teammate_counts.values():
        if count > 1:
            penalties += (count - 1) * 20

    for count in opponent_counts.values():
        if count > 2:
            penalties += (count - 2) * 10

    return 1000 - penalties

@app.route('/api/generate', methods=['POST'])
def generate_schedule():
    data = request.json
    event_id = data.get('eventId')
    selected_ids = data.get('selectedPlayers', [])
    drill_ids = data.get('drillPlayers', [])
    
    event = Event.query.get(event_id)
    if not event:
        return jsonify({'error': 'Event not found'}), 404
    
    selected_players = [Player.query.get(pid).to_dict() for pid in selected_ids]
    drill_players = [Player.query.get(pid).to_dict() for pid in drill_ids]
    
    settings = {
        'maxTeammates': int(Setting.get('maxTeammates', 1)),
        'maxOpponents': int(Setting.get('maxOpponents', 2))
    }
    
    # Generate schedule using algorithm
    best_schedule = None
    best_score = -999999

    for i in range(1):
        candidate = generate_schedule_algorithm(event.to_dict(), selected_players, drill_players, settings)
        candidate_score = calculate_schedule_score(candidate)

        if candidate_score > best_score:
            best_score = candidate_score
            best_schedule = candidate

    schedule_result = best_schedule
    schedule_result['optimizationScore'] = best_score

    # Ajoute les joueurs avec leur genre pour l'export Excel
    schedule_result['players'] = selected_players

    # Save to database
    schedule = Schedule(
        event_id=event_id,
        selected_players=json.dumps(selected_ids),
        drill_players=json.dumps(drill_ids),
        schedule_data=json.dumps(schedule_result)
    )
    db.session.add(schedule)
    db.session.commit()

    return jsonify(schedule_result)

@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    data = request.json
    if not data:
        return jsonify({'error': 'Aucune cédule à exporter'}), 400

    schedule = data
    event = schedule.get('event', {})
    periods = schedule.get('periods', [])

    def normalize_name(value):
        if not value:
            return ''
        return ' '.join(str(value).strip().upper().split())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cédule de la journée'

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    title_font = Font(bold=True, size=12)
    bold = Font(bold=True)

    grey_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
    light_fill = PatternFill(fill_type='solid', fgColor='F7F7F7')

    thin = Side(style='thin', color='A6A6A6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ==============================
    # Titre principal
    # ==============================

# ==============================
# En-tête style VBA - Cédule de la journée
# ==============================

    # Ligne 1 : informations générales
    ws.cell(row=1, column=3, value='Jeudi 2026-04-09 (Drill)')
    ws.cell(row=1, column=5, value='Ligue: DSP')
    ws.cell(row=1, column=7, value='Événement: Les Jeudis DSP')
    ws.cell(row=1, column=11, value='Endroit: Tennis 13')

    for cell_ref in ['C1', 'E1', 'G1', 'K1']:
        ws[cell_ref].font = Font(bold=True, size=11)
        ws[cell_ref].alignment = center

    # Ligne 2 : heures
    col = 3
    for period in periods:
        time_label = period.get('time', '')

        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)

        cell = ws.cell(row=2, column=col, value=time_label)
        cell.font = Font(bold=True, size=11)
        cell.alignment = center
        cell.border = border

        # Appliquer aussi la bordure à la cellule fusionnée de droite
        ws.cell(row=2, column=col + 1).border = border

        col += 2

    # Colonnes A/B sans titre visible, comme ton VBA
    ws.cell(row=2, column=1, value='')
    ws.cell(row=2, column=2, value='')

    # ==============================
    # Ligne 2 : Terrain / Côté + Heures
    # ==============================    

    ws.cell(row=2, column=1, value='Terrain')
    ws.cell(row=2, column=2, value='Côté')

    for c in [1, 2]:
        ws.cell(row=2, column=c).font = bold
        ws.cell(row=2, column=c).alignment = center
        ws.cell(row=2, column=c).fill = grey_fill
        ws.cell(row=2, column=c).border = border

    col = 3
    for period in periods:
        time_label = period.get('time', '')

        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)

        cell = ws.cell(row=2, column=col, value=time_label)
        cell.font = Font(bold=True, size=11)
        cell.alignment = center
        cell.border = border

        ws.cell(row=2, column=col + 1).border = border

        col += 2

        # Ajouter colonne séparation
        #ws.column_dimensions[get_column_letter(col)].width = 4

        #col += 1

    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.freeze_panes = 'C3'

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold = Font(bold=True)
    title_font = Font(bold=True, size=12)

    header_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
    light_fill = PatternFill(fill_type='solid', fgColor='F7F7F7')

    thin = Side(style='thin', color='A6A6A6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_stats = wb.create_sheet('Cédule - Statistiques')
    ws_partners = wb.create_sheet('ParrJoueursCoéquipiers')
    ws_opponents = wb.create_sheet('ParrJoueursAdversaires')

    ws_mixed = wb.create_sheet('Double Mixtes')

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    title_font = Font(bold=True, size=12)
    bold = Font(bold=True)

    grey_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
    light_fill = PatternFill(fill_type='solid', fgColor='F7F7F7')

    thin = Side(style='thin', color='A6A6A6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Déterminer le nombre maximal de terrains
    max_courts = 0
    for period in periods:
        courts = period.get('courts', [])
        if len(courts) > max_courts:
            max_courts = len(courts)

    # Ligne 1 : en-tête principal
    total_columns = 2 + (len(periods) * 2)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    title_font = Font(bold=True, size=12)
    bold = Font(bold=True)

    grey_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
    light_fill = PatternFill(fill_type='solid', fgColor='F2F2F2')

    thin = Side(style='thin', color='A6A6A6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Ligne 1 : en-tête principal
    total_columns = 2 + (len(periods) * 2)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)

    event_day = str(event.get('day', '')).strip()
    event_name = str(event.get('name', '')).strip()

    league_name = 'DSP'
    location_name = 'Tennis 13'

    header_text = f"{event_day} (Drill)    Ligue: {league_name}    Événement: {event_name}    Endroit: {location_name}"

    header_cell = ws.cell(row=1, column=1, value=header_text)
    header_cell.font = Font(bold=True, size=12)
    header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_cell.fill = grey_fill
    header_cell.border = border

    # Ligne 1 : en-tête principal
    total_columns = 2 + (len(periods) * 2)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)

    event_day = str(event.get('day', '')).strip()
    event_name = str(event.get('name', '')).strip()

    # Temporaire: date fixe pour reproduire le modèle VBA
    event_date = "2026-04-09"

    league_name = 'DSP'
    location_name = 'Tennis 13'

    header_day = f"{event_day} {event_date}".strip()
    header_text = f"{header_day} (Drill)    Ligue: {league_name}    Événement: {event_name}    Endroit: {location_name}"

    header_cell = ws.cell(row=1, column=1, value=header_text)
    header_cell.font = Font(bold=True, size=12)
    header_cell.alignment = center
    header_cell.fill = grey_fill
    header_cell.border = border

    # Ligne 2 : heures
    ws.cell(row=2, column=1, value='Terrain')
    ws.cell(row=2, column=2, value='Côté')

    for c in [1, 2]:
        ws.cell(row=2, column=c).font = bold
        ws.cell(row=2, column=c).alignment = center
        ws.cell(row=2, column=c).fill = grey_fill
        ws.cell(row=2, column=c).border = border

    col = 3
    for period in periods:
        time_label = period.get('time', '')
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        cell = ws.cell(row=2, column=col, value=time_label)
        cell.font = bold
        cell.alignment = center
        cell.fill = grey_fill
        cell.border = border

        ws.cell(row=2, column=col + 1).fill = grey_fill
        ws.cell(row=2, column=col + 1).border = border

        col += 2

    # Données : 2 lignes par terrain (A et B)
    row = 3

    terrain_numbers = [3, 4, 5, 6, 7, 8, 10, 11, 12, 99]

    for court_index in range(max_courts):

        fill_terrain = light_fill if court_index % 2 == 0 else PatternFill(fill_type='solid', fgColor='FFFFFF')
        terrain_no = terrain_numbers[court_index] if court_index < len(terrain_numbers) else court_index + 1

        # Fusion Terrain
        ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=1)

        terrain_cell = ws.cell(row=row, column=1, value=terrain_no)
        terrain_cell.font = bold
        terrain_cell.alignment = center
        terrain_cell.fill = fill_terrain
        terrain_cell.border = border

        # Colonne Côté
        ws.cell(row=row, column=2, value='A').alignment = center
        ws.cell(row=row+1, column=2, value='B').alignment = center

        ws.cell(row=row, column=2).border = border
        ws.cell(row=row+1, column=2).border = border
        ws.cell(row=row, column=2).fill = fill_terrain
        ws.cell(row=row+1, column=2).fill = fill_terrain

        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row+1, column=1).alignment = center

        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row+1, column=2).alignment = center

        col = 3

        for period in periods:
            courts = period.get('courts', [])

            if court_index < len(courts):
                court = courts[court_index]

                side_a = court.get('sideA', {})
                side_b = court.get('sideB', {})

                a1 = side_a.get('player1', {}).get('fullName', '')
                a2 = side_a.get('player2', {}).get('fullName', '')
                b1 = side_b.get('player1', {}).get('fullName', '')
                b2 = side_b.get('player2', {}).get('fullName', '')

                ws.cell(row=row, column=col, value=a1)
                ws.cell(row=row, column=col+1, value=a2)

                ws.cell(row=row+1, column=col, value=b1)
                ws.cell(row=row+1, column=col+1, value=b2)

                # Format
                for r in [row, row+1]:
                    for c in [col, col+1]:
                        ws.cell(row=r, column=c).alignment = left
                        ws.cell(row=r, column=c).border = border

            col += 2

        row += 2

        # ==============================
        # Bordure extérieure du tableau
        # ==============================

        thick = Side(style='medium', color='000000')

        max_col = 2 + len(periods)*2
        max_row = row - 1

        # Bordure gauche et droite
        for r in range(2, max_row+1):
            ws.cell(row=r, column=1).border = Border(left=thick)
            ws.cell(row=r, column=max_col).border = Border(right=thick)

        # Bordure haut et bas
        for c in range(1, max_col+1):
            ws.cell(row=2, column=c).border = Border(top=thick)
            ws.cell(row=max_row, column=c).border = Border(bottom=thick)

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 6

        col_idx = 3
        for period in periods:
            ws.column_dimensions[get_column_letter(col_idx)].width = 16
            ws.column_dimensions[get_column_letter(col_idx+1)].width = 16
            col_idx += 3  # 2 joueurs + 1 séparation

    # Largeur des colonnes
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 5

    for col_idx in range(3, 3 + len(periods)*2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 22

        for r in range(3, row):
            ws.row_dimensions[r].height = 28

    # Bordure entre terrains
    thick = Side(style='medium', color='000000')

    for r in range(3, row, 2):
        for c in range(1, 3 + len(periods)*2):
            ws.cell(row=r, column=c).border = Border(top=thick)

    # Alignement des noms
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

      # ==============================
    # Feuille : Cédule pour chaque joueur
    # ==============================

    ws_players = wb.create_sheet('Cédule pour chaque joueur')
    ws_players.sheet_view.showGridLines = False

    p_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    p_left = Alignment(horizontal='left', vertical='center', wrap_text=False)

    p_bold = Font(name='Calibri', size=11, bold=True)
    p_normal = Font(name='Calibri', size=11)

    p_gray = PatternFill(fill_type='solid', fgColor='D9D9D9')
    p_light = PatternFill(fill_type='solid', fgColor='F7F7F7')

    p_thin = Side(style='thin', color='A6A6A6')
    p_border = Border(left=p_thin, right=p_thin, top=p_thin, bottom=p_thin)

    # Liste des joueurs A:D
    headers_players = ['(F) Nom complet', 'Prénom', 'Nom', 'Courriel']

    for col_idx, header in enumerate(headers_players, start=1):
        cell = ws_players.cell(row=1, column=col_idx, value=header)
        cell.font = p_bold
        cell.alignment = p_center
        cell.fill = p_gray
        cell.border = p_border

    selected_players_list = schedule.get('players', [])

    for row_idx, player in enumerate(selected_players_list, start=2):
        full_name = player.get('fullName', '')
        first_name = player.get('firstName', '')
        last_name = player.get('lastName', '')
        email = player.get('email', '')

        values = [full_name, first_name, last_name, email]

        for col_idx, value in enumerate(values, start=1):
            cell = ws_players.cell(row=row_idx, column=col_idx, value=value)
            cell.font = p_normal
            cell.alignment = p_left
            cell.border = p_border

    # Construire la cédule individuelle
    player_schedule = {}

    for player in selected_players_list:
        name = player.get('fullName', '')
        if name:
            player_schedule[name] = []

    for period in periods:
        period_time = period.get('time', '')
        courts = period.get('courts', [])

        for court in courts:
            terrain = court.get('number', '')

            side_a = court.get('sideA', {})
            side_b = court.get('sideB', {})

            a1 = side_a.get('player1', {})
            a2 = side_a.get('player2', {})
            b1 = side_b.get('player1', {})
            b2 = side_b.get('player2', {})

            a1_name = a1.get('fullName', '')
            a2_name = a2.get('fullName', '')
            b1_name = b1.get('fullName', '')
            b2_name = b2.get('fullName', '')

            # Chaque joueur est toujours placé en première colonne de son équipe
            match_rows = [
                (a1_name, a2_name, b1_name, b2_name, terrain, period_time),
                (a2_name, a1_name, b1_name, b2_name, terrain, period_time),
                (b1_name, b2_name, a1_name, a2_name, terrain, period_time),
                (b2_name, b1_name, a1_name, a2_name, terrain, period_time),
            ]

            for joueur, partenaire, adv1, adv2, terrain_no, heure in match_rows:
                if joueur in player_schedule:
                    player_schedule[joueur].append([
                        joueur,
                        partenaire,
                        adv1,
                        adv2,
                        terrain_no,
                        heure
                    ])

    # Écrire les blocs de cédule
    block_height = 9
    matches_per_player = 7

    def write_player_block(start_row, start_col, player_name, matches):
        for idx in range(matches_per_player):
            row_num = start_row + idx

            if idx < len(matches):
                values = matches[idx]
            else:
                values = ['', '', '', '', '', '']

            for offset, value in enumerate(values):
                cell = ws_players.cell(row=row_num, column=start_col + offset, value=value)
                cell.font = p_normal
                cell.alignment = p_left if offset < 4 else p_center
                cell.border = p_border

        # Lignes de séparation après chaque joueur
        for spacer_row in [start_row + 7, start_row + 8]:
            for offset in range(6):
                cell = ws_players.cell(row=spacer_row, column=start_col + offset)
                cell.fill = p_light

    row_start = 2

    for idx, player in enumerate(selected_players_list):
        player_name = player.get('fullName', '')
        matches = player_schedule.get(player_name, [])

        block_row = row_start + (idx * block_height)

        # Zone principale E:J
        write_player_block(block_row, 5, player_name, matches)

        # Zone secondaire Q:V, comme dans le VBA
        #write_player_block(block_row, 17, player_name, matches)

    # Largeurs comme le modèle VBA
    widths = {
        'A': 21,
        'B': 13,
        'C': 13,
        'D': 28,
        'E': 21,
        'F': 13,
        'G': 13,
        'H': 13,
        'I': 5,
        'J': 12,
        'K': 13,
        'L': 13,
        'M': 13,
        'N': 13,
        'O': 13,
        'P': 13,
        'Q': 21,
        'R': 13,
        'S': 13,
        'T': 13,
        'U': 5,
        'V': 12
    }

    for col_letter, width in widths.items():
        ws_players.column_dimensions[col_letter].width = width

    # Cacher les colonnes A à D
    for col in ['A', 'B', 'C', 'D']:
        ws_players.column_dimensions[col].hidden = True

    # Hauteur lignes style VBA
    max_row_players = row_start + (len(selected_players_list) * block_height)

    for r in range(1, max_row_players + 1):
        ws_players.row_dimensions[r].height = 13.5

    ws_players.freeze_panes = 'E2'

    # ==============================
    # Feuille : Cédule - Statistiques
    # ==============================
    ws_stats['A4'] = 'Jour et Date de la rencontre'
    ws_stats['A4'].font = Font(bold=True, size=14)

    ws_stats['K2'] = 'Commentaires'
    ws_stats['K2'].font = Font(bold=True, size=14)
    ws_stats['K2'].alignment = center

    # ===== Données =====
    event_day = str(event.get('day', '')).strip()
    event_name = str(event.get('name', '')).strip()
    event_date = '4/9/2026'  # temporaire

    selected_players_set = set()
    drill_players_set = set()
    used_courts = []

    for period in periods:
        period_name = period.get('name', '')
        courts = period.get('courts', [])

        for court in courts:
            terrain = court.get('number', '')

            if terrain not in used_courts:
                used_courts.append(terrain)

            side_a = court.get('sideA', {})
            side_b = court.get('sideB', {})

            a1 = side_a.get('player1', {}).get('fullName', '')
            a2 = side_a.get('player2', {}).get('fullName', '')
            b1 = side_b.get('player1', {}).get('fullName', '')
            b2 = side_b.get('player2', {}).get('fullName', '')

            for p in [a1, a2, b1, b2]:
                if p:
                    selected_players_set.add(p)

            if 'Drill' in period_name:
                for p in [a1, a2, b1, b2]:
                    if p:
                        drill_players_set.add(p)

    used_courts = sorted(used_courts, key=lambda x: int(x) if str(x).isdigit() else 999)
    #used_courts.sort()

    selected_count = len(selected_players_set)
    drill_count = len(drill_players_set)
    drill_yes_no = 'Oui' if drill_count > 0 else 'Non'

    # Heure sans secondes
    start_time = str(event.get('startTime', '')).split(':')
    end_time = str(event.get('endTime', '')).split(':')

    start_time = f"{start_time[0]}:{start_time[1]}" if len(start_time) >= 2 else ''
    end_time = f"{end_time[0]}:{end_time[1]}" if len(end_time) >= 2 else ''

    time_range = f"{start_time} - {end_time}"

    drill_minutes = event.get('drillMinutes', 60)
    period_duration = event.get('periodDuration', 20)

    # ===== Bloc 1 =====
    ws_stats['A6'] = 'Nom de la journée'
    ws_stats['A7'] = 'Date de la rencontre'

    ws_stats['I6'] = event_day
    ws_stats['I7'] = event_date

    # ===== Bloc 2 =====
    ws_stats['A10'] = 'Nombre de joueurs sélectionnés'
    ws_stats['I10'] = selected_count

    # ===== Bloc 3 =====
    ws_stats['A13'] = 'Drill et Nombre de joueurs pour la Drill'
    ws_stats['A14'] = 'Oui / Non'
    ws_stats['A15'] = 'Nombre de joueurs'
    ws_stats['A16'] = 'Durée en minutes'

    ws_stats['I14'] = drill_yes_no
    ws_stats['I15'] = drill_count
    ws_stats['I16'] = drill_minutes

    # ===== Bloc 4 =====
    ws_stats['A19'] = 'Horaire et parties'
    ws_stats['A20'] = 'Heure début et fin'
    ws_stats['A21'] = "Durée d'une partie en minutes"

    ws_stats['I20'] = time_range
    ws_stats['I21'] = period_duration

    # ===== Bloc 5 =====
    ws_stats['A24'] = 'Nombre de terrains réservés'
    ws_stats['I24'] = len(used_courts)

    ws_stats['A25'] = 'Numéros'

    col_start = 10  # colonne J
    for i, terrain in enumerate(used_courts):
        ws_stats.cell(row=25, column=col_start + i, value=terrain)

    # ===== Commentaires =====
    ws_stats.merge_cells('K3:N18')
    ws_stats['K3'] = ''

    thin = Side(style='thin', color='A6A6A6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws_stats['K3'].border = border

    # ===== Styles =====
    bold_big = Font(bold=True, size=14)
    bold = Font(bold=True)

    for cell in ['A4', 'A10', 'A13', 'A19', 'A24']:
        ws_stats[cell].font = bold_big

    for cell in ['A6', 'A7', 'A14', 'A15', 'A16', 'A20', 'A21', 'A25']:
        ws_stats[cell].font = bold

    for row in range(1, 35):
        ws_stats.row_dimensions[row].height = 22

    # Largeurs
    ws_stats.column_dimensions['A'].width = 42
    ws_stats.column_dimensions['I'].width = 22

    for col in ['J', 'K', 'L', 'M', 'N']:
        ws_stats.column_dimensions[col].width = 10

    filename = "Cedule_de_la_journee.xlsx"

    # ==============================
    # Feuille : ParrJoueursCoéquipiers
    # ==============================

    ws_partners['A1'] = 'ParrJoueursCoéquipiers'
    ws_partners['A1'].font = Font(bold=True, size=14)
    ws_partners['A1'].alignment = center

    headers = ['Joueur', 'Partenaire', 'Nb fois']

    for col_idx, header in enumerate(headers, start=1):
        cell = ws_partners.cell(row=2, column=col_idx, value=header)
        cell.font = bold
        cell.alignment = center
        cell.fill = grey_fill
        cell.border = border

    # Dictionnaire : (joueur, partenaire) → nb
    partner_counts = {}

    for period in periods:
        courts = period.get('courts', [])

        for court in courts:
            side_a = court.get('sideA', {})
            side_b = court.get('sideB', {})

            a1 = side_a.get('player1', {}).get('fullName', '')
            a2 = side_a.get('player2', {}).get('fullName', '')
            b1 = side_b.get('player1', {}).get('fullName', '')
            b2 = side_b.get('player2', {}).get('fullName', '')

            pairs = [
                (a1, a2),
                (a2, a1),
                (b1, b2),
                (b2, b1)
            ]

            for joueur, partenaire in pairs:
                if joueur and partenaire:
                    key = (joueur, partenaire)
                    partner_counts[key] = partner_counts.get(key, 0) + 1

    # Trier comme VBA :
    # 1. Nb de fois décroissant
    # 2. Nom du joueur A-Z
    # 3. Nom du coéquipier A-Z
    sorted_pairs = sorted(
        partner_counts.items(),
        key=lambda x: (-x[1], x[0][0], x[0][1])
    )

    # Headers
    headers = ['Nom du joueur', 'Nom du coéquipier', 'Nb de fois', 'Type Équipe']

    for col_idx, header in enumerate(headers, start=1):
        cell = ws_partners.cell(row=2, column=col_idx, value=header)
        cell.font = bold
        cell.alignment = center
        cell.fill = grey_fill
        cell.border = border

    row_p = 3

    for (joueur, partenaire), count in sorted_pairs:
        gender_joueur = player_gender_map.get(normalize_name(joueur), '')
        gender_partenaire = player_gender_map.get(normalize_name(partenaire), '')

        if gender_joueur and gender_partenaire:
            if gender_joueur != gender_partenaire:
                type_equipe = 'MX'
            elif gender_joueur == 'F':
                type_equipe = 'DF'
            else:
                type_equipe = 'DM'
        else:
            type_equipe = ''

        values = [
            joueur,
            partenaire,
            count,
            type_equipe
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws_partners.cell(row=row_p, column=col_idx, value=value)
            cell.alignment = center
            cell.border = border

        row_p += 1

    # Largeurs colonnes
    ws_partners.column_dimensions['A'].width = 30
    ws_partners.column_dimensions['B'].width = 30
    ws_partners.column_dimensions['C'].width = 12
    ws_partners.column_dimensions['D'].width = 14

    row_p = 3
    for (joueur, partenaire), count in sorted_pairs:
        ws_partners.cell(row=row_p, column=1, value=joueur).alignment = center
        ws_partners.cell(row=row_p, column=2, value=partenaire).alignment = center
        ws_partners.cell(row=row_p, column=3, value=count).alignment = center

        for c in range(1, 4):
            ws_partners.cell(row=row_p, column=c).border = border

        row_p += 1

    # Largeurs colonnes
    ws_partners.column_dimensions['A'].width = 30
    ws_partners.column_dimensions['B'].width = 30
    ws_partners.column_dimensions['C'].width = 10

    # Hauteur lignes
    ws_partners.row_dimensions[1].height = 24
    ws_partners.row_dimensions[2].height = 22

    for r in range(3, row_p):
        ws_partners.row_dimensions[r].height = 20

    # ==============================
    # Feuille : ParrJoueursAdversaires
    # ==============================

    ws_opponents['A1'] = 'ParrJoueursAdversaires'
    ws_opponents['A1'].font = Font(bold=True, size=14)
    ws_opponents['A1'].alignment = center

    headers_opp = ['Joueur', 'Adversaire', 'Nb fois']

    for col_idx, header in enumerate(headers_opp, start=1):
        cell = ws_opponents.cell(row=2, column=col_idx, value=header)
        cell.font = bold
        cell.alignment = center
        cell.fill = grey_fill
        cell.border = border

    # Dictionnaire : (joueur, adversaire) -> nb
    opponent_counts = {}

    for period in periods:
        courts = period.get('courts', [])

        for court in courts:
            side_a = court.get('sideA', {})
            side_b = court.get('sideB', {})

            a1 = side_a.get('player1', {}).get('fullName', '')
            a2 = side_a.get('player2', {}).get('fullName', '')
            b1 = side_b.get('player1', {}).get('fullName', '')
            b2 = side_b.get('player2', {}).get('fullName', '')

            matchups = [
                (a1, b1), (a1, b2),
                (a2, b1), (a2, b2),
                (b1, a1), (b1, a2),
                (b2, a1), (b2, a2),
            ]

            for joueur, adversaire in matchups:
                if joueur and adversaire:
                    key = (joueur, adversaire)
                    opponent_counts[key] = opponent_counts.get(key, 0) + 1

    # Trier par joueur puis adversaire
    sorted_opponents = sorted(opponent_counts.items(), key=lambda x: (x[0][0], x[0][1]))

    row_o = 3
    for (joueur, adversaire), count in sorted_opponents:
        ws_opponents.cell(row=row_o, column=1, value=joueur).alignment = center
        ws_opponents.cell(row=row_o, column=2, value=adversaire).alignment = center
        ws_opponents.cell(row=row_o, column=3, value=count).alignment = center

        for c in range(1, 4):
            ws_opponents.cell(row=row_o, column=c).border = border

        row_o += 1

    # Largeurs colonnes
    ws_opponents.column_dimensions['A'].width = 30
    ws_opponents.column_dimensions['B'].width = 30
    ws_opponents.column_dimensions['C'].width = 10

    # Hauteur lignes
    ws_opponents.row_dimensions[1].height = 24
    ws_opponents.row_dimensions[2].height = 22

    for r in range(3, row_o):
        ws_opponents.row_dimensions[r].height = 20

    # Dictionnaire nom complet -> genre
    player_gender_map = {}

    for p in schedule.get('players', []):
        full_name = normalize_name(p.get('fullName', ''))
        gender = str(p.get('gender', '')).strip().upper()

        if full_name and gender:
            player_gender_map[full_name] = gender

    # ==============================
    # Feuille : Double Mixtes
    # ==============================

    ws_mixed['A1'] = 'Double Mixtes'
    ws_mixed['A1'].font = Font(bold=True, size=14)
    ws_mixed['A1'].alignment = center

    headers = [
        'Période',
        'Heure',
        'Terrain',
        'Équipe A',
        'Mixte A',
        'Équipe B',
        'Mixte B'
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws_mixed.cell(row=2, column=col_idx, value=header)
        cell.font = bold
        cell.alignment = center
        cell.fill = grey_fill
        cell.border = border

    row_m = 3

    for period in periods:
        period_name = period.get('name', '')
        period_time = period.get('time', '')
        courts = period.get('courts', [])

        for court in courts:
            terrain = court.get('number', '')

            side_a = court.get('sideA', {})
            side_b = court.get('sideB', {})

            a1 = side_a.get('player1', {})
            a2 = side_a.get('player2', {})
            b1 = side_b.get('player1', {})
            b2 = side_b.get('player2', {})

            # Noms
            team_a = f"{a1.get('fullName', '')} / {a2.get('fullName', '')}"
            team_b = f"{b1.get('fullName', '')} / {b2.get('fullName', '')}"

            # Mixte
            gender_a1 = player_gender_map.get(normalize_name(a1.get('fullName', '')), '')
            gender_a2 = player_gender_map.get(normalize_name(a2.get('fullName', '')), '')
            gender_b1 = player_gender_map.get(normalize_name(b1.get('fullName', '')), '')
            gender_b2 = player_gender_map.get(normalize_name(b2.get('fullName', '')), '')

            mix_a = 'Oui' if gender_a1 and gender_a2 and gender_a1 != gender_a2 else 'Non'
            mix_b = 'Oui' if gender_b1 and gender_b2 and gender_b1 != gender_b2 else 'Non'

            values = [
                period_name,
                period_time,
                terrain,
                team_a,
                mix_a,
                team_b,
                mix_b
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws_mixed.cell(row=row_m, column=col_idx, value=value)
                cell.alignment = center
                cell.border = border

            row_m += 1

    # Largeurs colonnes
    ws_mixed.column_dimensions['A'].width = 18
    ws_mixed.column_dimensions['B'].width = 12
    ws_mixed.column_dimensions['C'].width = 10
    ws_mixed.column_dimensions['D'].width = 32
    ws_mixed.column_dimensions['E'].width = 10
    ws_mixed.column_dimensions['F'].width = 32
    ws_mixed.column_dimensions['G'].width = 10

    # Hauteur lignes
    ws_mixed.row_dimensions[1].height = 24
    ws_mixed.row_dimensions[2].height = 22

    for r in range(3, row_m):
        ws_mixed.row_dimensions[r].height = 20

     # ==============================
    # Feuille : Cédule de la journée (V2) - Style original VBA
    # ==============================

    ws_day_v2 = wb.create_sheet('Cédule de la journée (V2)')
    ws_day_v2.sheet_view.showGridLines = False
    ws_day_v2.page_setup.orientation = 'landscape'
    ws_day_v2.page_setup.fitToWidth = 1
    ws_day_v2.page_setup.fitToHeight = 0
    ws_day_v2.sheet_properties.pageSetUpPr.fitToPage = True

    v2_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    v2_left = Alignment(horizontal='left', vertical='center')

    v2_bold = Font(name='Calibri', size=11, bold=True)
    v2_normal = Font(name='Calibri', size=11)

    v2_gray = PatternFill(fill_type='solid', fgColor='D9D9D9')
    v2_white = PatternFill(fill_type='solid', fgColor='FFFFFF')

    v2_thin = Side(style='thin', color='A6A6A6')
    v2_border = Border(left=v2_thin, right=v2_thin, top=v2_thin, bottom=v2_thin)

    terrain_numbers_v2 = [3, 4, 5, 6, 7, 8, 10, 11, 12, 99]
    max_courts_v2 = max(len(period.get('courts', [])) for period in periods) if periods else 0

    row_v2 = 1
    courts_per_block = 4

    def write_v2_header(row_number):
        col = 1
        match_no = 1

        for period in periods:
            time_label = period.get('time', '')
            header_text = f"{time_label} Match {match_no}"

            if 'Drill' in period.get('name', ''):
                header_text = f"{time_label} (Drill) Match {match_no}"

            ws_day_v2.merge_cells(
                start_row=row_number,
                start_column=col,
                end_row=row_number,
                end_column=col + 1
            )

            cell = ws_day_v2.cell(row=row_number, column=col, value=header_text)
            cell.font = v2_bold
            cell.alignment = v2_center
            cell.fill = v2_gray
            cell.border = v2_border

            ws_day_v2.cell(row=row_number, column=col + 1).fill = v2_gray
            ws_day_v2.cell(row=row_number, column=col + 1).border = v2_border

            # colonne séparatrice
            ws_day_v2.column_dimensions[get_column_letter(col + 2)].width = 2.5

            col += 3
            match_no += 1

    for block_start in range(0, max_courts_v2, courts_per_block):
        write_v2_header(row_v2)
        row_v2 += 1

        for court_index in range(block_start, min(block_start + courts_per_block, max_courts_v2)):
            terrain_no = terrain_numbers_v2[court_index] if court_index < len(terrain_numbers_v2) else court_index + 1

            # Ligne Terrain
            col = 1
            for period in periods:
                ws_day_v2.merge_cells(
                    start_row=row_v2,
                    start_column=col,
                    end_row=row_v2,
                    end_column=col + 1
                )

                cell = ws_day_v2.cell(row=row_v2, column=col, value=f"Terrain {terrain_no}")
                cell.font = Font(name='Calibri', size=11, bold=False)  # 👈 moins agressif
                cell.alignment = v2_center
                cell.fill = PatternFill(fill_type='solid', fgColor='F7F7F7')  # 👈 plus léger
                cell.border = v2_border

                ws_day_v2.cell(row=row_v2, column=col + 1).fill = v2_gray
                ws_day_v2.cell(row=row_v2, column=col + 1).border = v2_border

                col += 3

            row_v2 += 1

            # Ligne Team A / Team B
            col = 1
            for period in periods:
                ws_day_v2.cell(row=row_v2, column=col, value='Team A')
                ws_day_v2.cell(row=row_v2, column=col + 1, value='Team B')

                for c in [col, col + 1]:
                    ws_day_v2.cell(row=row_v2, column=c).font = v2_bold
                    ws_day_v2.cell(row=row_v2, column=c).alignment = v2_center
                    ws_day_v2.cell(row=row_v2, column=c).fill = PatternFill(fill_type='solid', fgColor='EFEFEF')
                    ws_day_v2.cell(row=row_v2, column=c).border = v2_border

                col += 3

            row_v2 += 1

            # Deux lignes de joueurs
            for player_line in [1, 2]:
                col = 1

                for period in periods:
                    courts = period.get('courts', [])

                    if court_index < len(courts):
                        court = courts[court_index]

                        side_a = court.get('sideA', {})
                        side_b = court.get('sideB', {})

                        if player_line == 1:
                            value_a = side_a.get('player1', {}).get('fullName', '')
                            value_b = side_b.get('player1', {}).get('fullName', '')
                        else:
                            value_a = side_a.get('player2', {}).get('fullName', '')
                            value_b = side_b.get('player2', {}).get('fullName', '')

                        ws_day_v2.cell(row=row_v2, column=col, value=value_a)
                        ws_day_v2.cell(row=row_v2, column=col + 1, value=value_b)

                        for c in [col, col + 1]:
                            ws_day_v2.cell(row=row_v2, column=c).font = v2_normal
                            ws_day_v2.cell(row=row_v2, column=c).alignment = v2_left
                            ws_day_v2.cell(row=row_v2, column=c).border = v2_border

                    col += 3

                row_v2 += 1

        # espace entre les blocs de terrains
        row_v2 += 2

    # Bloc Assis
    write_v2_header(row_v2)
    row_v2 += 1

    col = 1
    for period in periods:
        ws_day_v2.merge_cells(
            start_row=row_v2,
            start_column=col,
            end_row=row_v2,
            end_column=col + 1
        )

        cell = ws_day_v2.cell(row=row_v2, column=col, value='Assis')
        cell.font = v2_bold
        cell.alignment = v2_center
        cell.fill = PatternFill(fill_type='solid', fgColor='EFEFEF')
        cell.border = v2_border

        sitting = period.get('sitting', [])
        names = [p.get('fullName', '') for p in sitting if p.get('fullName', '')]

        ws_day_v2.cell(row=row_v2 + 1, column=col, value=', '.join(names))
        ws_day_v2.merge_cells(
            start_row=row_v2 + 1,
            start_column=col,
            end_row=row_v2 + 2,
            end_column=col + 1
        )

        ws_day_v2.cell(row=row_v2 + 1, column=col).alignment = v2_left
        ws_day_v2.cell(row=row_v2 + 1, column=col).border = v2_border

        col += 3

    # Largeurs
    col = 1
    for period in periods:
        ws_day_v2.column_dimensions[get_column_letter(col)].width = 18
        ws_day_v2.column_dimensions[get_column_letter(col + 1)].width = 18
        ws_day_v2.column_dimensions[get_column_letter(col + 2)].width = 3
        col += 3

    # Hauteurs
    for r in range(1, row_v2 + 3):
        ws_day_v2.row_dimensions[r].height = 18

    ws_day_v2.freeze_panes = 'A2'
    #ws_day_v2.print_area = f"A1:{get_column_letter(3 * len(periods))}{row_v2 + 2}"

    # ==============================
    # Feuille : Statistiques avancées
    # ==============================

    ws_stats_adv = wb.create_sheet('Statistiques avancées')

    headers = [
        'Joueur',
        'Matchs',
        'Partenaires uniques',
        'Adversaires uniques',
        'Côté A',
        'Côté B',
        'Doubles mixtes',
        'Drill',
        'Score équité',
        'Évaluation'
    ]

    # En-têtes en ligne 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_stats_adv.cell(row=3, column=col_idx, value=header)
        cell.font = bold
        cell.alignment = center
        cell.fill = grey_fill
        cell.border = border

    # Préparer structures
    stats = {}

    for p in schedule.get('players', []):
        name = p.get('fullName', '')
        stats[name] = {
            'matches': 0,
            'partners': set(),
            'opponents': set(),
            'sideA': 0,
            'sideB': 0,
            'mixed': 0,
            'drill': 1 if p.get('drill') else 0
        }

    # Parcours des matchs
    for period in periods:
        courts = period.get('courts', [])

        for court in courts:
            side_a = court.get('sideA', {})
            side_b = court.get('sideB', {})

            a1 = side_a.get('player1', {})
            a2 = side_a.get('player2', {})
            b1 = side_b.get('player1', {})
            b2 = side_b.get('player2', {})

            players = [
                (a1, 'A', a2, [b1, b2]),
                (a2, 'A', a1, [b1, b2]),
                (b1, 'B', b2, [a1, a2]),
                (b2, 'B', b1, [a1, a2])
            ]

            for player, side, partner, opponents in players:
                name = player.get('fullName', '')
                if name not in stats:
                    continue

                stats[name]['matches'] += 1
                stats[name]['partners'].add(partner.get('fullName', ''))

                for opp in opponents:
                    stats[name]['opponents'].add(opp.get('fullName', ''))

                if side == 'A':
                    stats[name]['sideA'] += 1
                else:
                    stats[name]['sideB'] += 1

                gender = player_gender_map.get(normalize_name(name), '')
                partner_gender = player_gender_map.get(normalize_name(partner.get('fullName', '')), '')

                if gender and partner_gender and gender != partner_gender:
                    stats[name]['mixed'] += 1

    # Trier par score d'équité décroissant
    sorted_stats = sorted(
        stats.items(),
        key=lambda item: len(item[1]['partners']) + len(item[1]['opponents']),
        reverse=True
    )

    # Écriture Excel à partir de la ligne 4
    row = 4

    for name, s in sorted_stats:
        score = len(s['partners']) + len(s['opponents'])

        if score < 6:
            evaluation = 'Faible'
        elif score < 10:
            evaluation = 'Moyen'
        else:
            evaluation = 'Bon'

        values = [
            name,
            s['matches'],
            len(s['partners']),
            len(s['opponents']),
            s['sideA'],
            s['sideB'],
            s['mixed'],
            s['drill'],
            score,
            evaluation
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws_stats_adv.cell(row=row, column=col_idx, value=value)
            cell.alignment = center
            cell.border = border

        row += 1

    # Coloration du score d'équité
    for r in range(4, row):
        score_cell = ws_stats_adv.cell(row=r, column=9)

        if score_cell.value is not None:
            if score_cell.value < 6:
                score_cell.fill = PatternFill(fill_type='solid', fgColor='FFC7CE')
            elif score_cell.value < 10:
                score_cell.fill = PatternFill(fill_type='solid', fgColor='FFEB9C')
            else:
                score_cell.fill = PatternFill(fill_type='solid', fgColor='C6EFCE')

    # Coloration de la colonne Évaluation
    for r in range(4, row):
        eval_cell = ws_stats_adv.cell(row=r, column=10)

        if eval_cell.value == 'Faible':
            eval_cell.fill = PatternFill(fill_type='solid', fgColor='FFC7CE')
        elif eval_cell.value == 'Moyen':
            eval_cell.fill = PatternFill(fill_type='solid', fgColor='FFEB9C')
        elif eval_cell.value == 'Bon':
            eval_cell.fill = PatternFill(fill_type='solid', fgColor='C6EFCE')

    # Alerte automatique
    nb_faible = 0

    for r in range(4, row):
        if ws_stats_adv.cell(row=r, column=10).value == 'Faible':
            nb_faible += 1

    if nb_faible > 0:
        message = f"⚠️ Déséquilibre détecté : {nb_faible} joueurs Faible"
        fill_color = 'FFC7CE'
    else:
        message = "✅ Cédule équilibrée"
        fill_color = 'C6EFCE'

    # Ligne 1 : alerte
    ws_stats_adv.merge_cells('A1:J1')
    ws_stats_adv['A1'] = message
    ws_stats_adv['A1'].font = Font(bold=True, size=14)
    ws_stats_adv['A1'].alignment = center
    ws_stats_adv['A1'].fill = PatternFill(fill_type='solid', fgColor=fill_color)
    ws_stats_adv['A1'].border = border

    # Ligne 2 : titre + score moyen
    ws_stats_adv.merge_cells('A2:H2')
    ws_stats_adv['A2'] = 'Statistiques avancées'
    ws_stats_adv['A2'].font = Font(bold=True, size=14)
    ws_stats_adv['A2'].alignment = center
    ws_stats_adv['A2'].fill = grey_fill
    ws_stats_adv['A2'].border = border

    ws_stats_adv['I2'] = 'Score moyen'
    ws_stats_adv['I2'].font = Font(bold=True)
    ws_stats_adv['I2'].alignment = center
    ws_stats_adv['I2'].fill = grey_fill
    ws_stats_adv['I2'].border = border

    ws_stats_adv['J2'] = f"=AVERAGE(I4:I{row-1})"
    ws_stats_adv['J2'].alignment = center
    ws_stats_adv['J2'].fill = PatternFill(fill_type='solid', fgColor='D9EAD3')
    ws_stats_adv['J2'].border = border

    ws_stats_adv['K1'] = 'Score optimisation'
    ws_stats_adv['K1'].font = Font(bold=True)
    ws_stats_adv['K1'].alignment = center
    ws_stats_adv['K1'].fill = grey_fill
    ws_stats_adv['K1'].border = border

    ws_stats_adv['K2'] = schedule.get('optimizationScore', '')
    ws_stats_adv['K2'].alignment = center
    ws_stats_adv['K2'].fill = PatternFill(fill_type='solid', fgColor='D9EAD3')
    ws_stats_adv['K2'].border = border

    ws_stats_adv.column_dimensions['K'].width = 18

    # Largeurs
    ws_stats_adv.column_dimensions['A'].width = 28
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws_stats_adv.column_dimensions[col].width = 18

    ws_stats_adv.freeze_panes = 'A4'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )    

@app.route('/api/export/<int:schedule_id>', methods=['GET'])
def export_schedule(schedule_id):    

    schedule = Schedule.query.get(schedule_id)
    if not schedule:
        return jsonify({'error': 'Schedule not found'}), 404
    
    schedule_data = json.loads(schedule.schedule_data)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Add sheets
    create_schedule_sheet(wb, schedule_data)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Pickleball_{schedule_data['event']['day']}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# ==================== ALGORITHM ====================

def generate_schedule_algorithm(event, all_players, drill_players, settings):
    """Main schedule generation with backtracking algorithm"""
    periods = []
    pairings = {
        'teammates': {p['id']: {} for p in all_players},
        'opponents': {p['id']: {} for p in all_players}
    }
    
    # Parse times
    start_minutes = time_to_minutes(event['startTime'])
    current_minutes = start_minutes
    
    # Drill period
    if event['drillMinutes'] > 0 and len(drill_players) >= 4:
        courts = generate_courts_with_constraints(drill_players, pairings, settings, True)
        periods.append({
            'name': '⚙️ Drill',
            'time': minutes_to_time(current_minutes),
            'isDrill': True,
            'courts': courts,
            'sitting': []
        })
        current_minutes += event['drillMinutes']
    
    # Regular periods
    end_minutes = time_to_minutes(event['endTime'])
    total_duration = end_minutes - current_minutes
    num_periods = total_duration // event['periodDuration']
    
    # Track play count
    play_count = {p['id']: 0 for p in all_players}
    
    for i in range(num_periods):
        # Sort by play count
        sorted_players = sorted(all_players, key=lambda p: (play_count[p['id']], -p['level']))
        
        max_playing = (len(all_players) // 4) * 4
        playing = sorted_players[:max_playing]
        sitting = sorted_players[max_playing:]
        
        courts = generate_courts_with_constraints(playing, pairings, settings, False)
        
        # Update play count
        for p in playing:
            play_count[p['id']] += 1
        
        periods.append({
            'name': f'Période {i + 1}',
            'time': minutes_to_time(current_minutes),
            'isDrill': False,
            'courts': courts,
            'sitting': sitting
        })
        current_minutes += event['periodDuration']
    
    return {
        'event': event,
        'players': all_players,
        'drillPlayers': drill_players,
        'periods': periods,
        'pairings': pairings,
        'timestamp': datetime.utcnow().isoformat()
    }

def generate_courts_with_constraints(players, pairings, settings, is_drill):
    """Generate courts using backtracking algorithm"""
    court_numbers = ['3', '4', '5', '6', '7', '8', '10', '11', '12']
    num_courts = len(players) // 4
    
    max_teammates = settings['maxTeammates']
    max_opponents = settings['maxOpponents']
    
    assignments = []
    
    def can_assign(p1, p2, p3, p4):
        tm12 = get_pairing_count(pairings['teammates'], p1['id'], p2['id'])
        tm34 = get_pairing_count(pairings['teammates'], p3['id'], p4['id'])
        
        if tm12 >= max_teammates or tm34 >= max_teammates:
            return False
        
        opp13 = get_pairing_count(pairings['opponents'], p1['id'], p3['id'])
        opp14 = get_pairing_count(pairings['opponents'], p1['id'], p4['id'])
        opp23 = get_pairing_count(pairings['opponents'], p2['id'], p3['id'])
        opp24 = get_pairing_count(pairings['opponents'], p2['id'], p4['id'])
        
        if opp13 >= max_opponents or opp14 >= max_opponents or \
           opp23 >= max_opponents or opp24 >= max_opponents:
            return False
        
        return True
    
    def score_match(p1, p2, p3, p4):
        tm12 = get_pairing_count(pairings['teammates'], p1['id'], p2['id'])
        tm34 = get_pairing_count(pairings['teammates'], p3['id'], p4['id'])
        opp_sum = sum([
            get_pairing_count(pairings['opponents'], p1['id'], p3['id']),
            get_pairing_count(pairings['opponents'], p1['id'], p4['id']),
            get_pairing_count(pairings['opponents'], p2['id'], p3['id']),
            get_pairing_count(pairings['opponents'], p2['id'], p4['id'])
        ])
        
        pairing_penalty = (tm12 + tm34) * 1000 + opp_sum * 100
        
        level_diff12 = abs(p1['level'] - p2['level'])
        level_diff34 = abs(p3['level'] - p4['level'])
        team_avg1 = (p1['level'] + p2['level']) / 2
        team_avg2 = (p3['level'] + p4['level']) / 2
        team_diff = abs(team_avg1 - team_avg2)
        
        level_penalty = (level_diff12 + level_diff34 + team_diff * 2) * 10
        
        return -(pairing_penalty + level_penalty)
    
    def backtrack(court_idx, remaining):
        if court_idx >= num_courts or len(remaining) < 4:
            return True
        
        n = len(remaining)
        candidates = []
        
        # Generate combinations
        for i in range(min(n - 3, 10)):
            for j in range(i + 1, min(n - 2, i + 10)):
                for k in range(j + 1, min(n - 1, j + 10)):
                    for l in range(k + 1, min(n, k + 10)):
                        p1, p2, p3, p4 = remaining[i], remaining[j], remaining[k], remaining[l]
                        
                        if can_assign(p1, p2, p3, p4):
                            candidates.append({
                                'players': [p1, p2, p3, p4],
                                'score': score_match(p1, p2, p3, p4)
                            })
        
        if not candidates:
            return False
        
        candidates.sort(key=lambda x: x['score'], reverse=True)
        
        for c in candidates[:5]:
            p1, p2, p3, p4 = c['players']
            
            assignments.append({'p1': p1, 'p2': p2, 'p3': p3, 'p4': p4, 'court': court_numbers[court_idx]})
            
            record_pairing(pairings['teammates'], p1['id'], p2['id'])
            record_pairing(pairings['teammates'], p3['id'], p4['id'])
            record_pairing(pairings['opponents'], p1['id'], p3['id'])
            record_pairing(pairings['opponents'], p1['id'], p4['id'])
            record_pairing(pairings['opponents'], p2['id'], p3['id'])
            record_pairing(pairings['opponents'], p2['id'], p4['id'])
            
            new_remaining = [p for p in remaining if p not in [p1, p2, p3, p4]]
            
            if backtrack(court_idx + 1, new_remaining):
                return True
            
            decrement_pairing(pairings['teammates'], p1['id'], p2['id'])
            decrement_pairing(pairings['teammates'], p3['id'], p4['id'])
            decrement_pairing(pairings['opponents'], p1['id'], p3['id'])
            decrement_pairing(pairings['opponents'], p1['id'], p4['id'])
            decrement_pairing(pairings['opponents'], p2['id'], p3['id'])
            decrement_pairing(pairings['opponents'], p2['id'], p4['id'])
            
            assignments.pop()
        
        return False
    
    backtrack(0, players)
    
    courts = []
    for a in assignments:
        courts.append({
            'number': a['court'],
            'sideA': {'player1': a['p1'], 'player2': a['p2']},
            'sideB': {'player1': a['p3'], 'player2': a['p4']}
        })
    
    return courts

def get_pairing_count(pairing_obj, id1, id2):
    return pairing_obj.get(id1, {}).get(id2, 0)

def record_pairing(pairing_obj, id1, id2):
    if id1 not in pairing_obj:
        pairing_obj[id1] = {}
    if id2 not in pairing_obj:
        pairing_obj[id2] = {}
    
    pairing_obj[id1][id2] = pairing_obj[id1].get(id2, 0) + 1
    pairing_obj[id2][id1] = pairing_obj[id2].get(id1, 0) + 1

def decrement_pairing(pairing_obj, id1, id2):
    if id1 in pairing_obj and id2 in pairing_obj[id1]:
        pairing_obj[id1][id2] -= 1
        if pairing_obj[id1][id2] <= 0:
            del pairing_obj[id1][id2]
    
    if id2 in pairing_obj and id1 in pairing_obj[id2]:
        pairing_obj[id2][id1] -= 1
        if pairing_obj[id2][id1] <= 0:
            del pairing_obj[id2][id1]

def time_to_minutes(time_str):
    parts = time_str.split(':')
    return int(parts[0]) * 60 + int(parts[1])

def minutes_to_time(minutes):
    h = minutes // 60
    m = minutes % 60
    return f"{h:02d}:{m:02d}"

def create_schedule_sheet(wb, schedule_data):
    ws = wb.create_sheet('Cédule de la journée')
    
    # Headers
    ws['A1'] = 'Terrain'
    ws['B1'] = 'Côté'
    col = 3
    
    for period in schedule_data['periods']:
        ws.cell(1, col, period['name'])
        ws.cell(1, col + 1, '')
        col += 2
    
    # Data
    row = 2
    for period in schedule_data['periods']:
        for court in period['courts']:
            # Côté A
            ws.cell(row, 1, court['number'])
            ws.cell(row, 2, 'A')
            ws.cell(row, 3, court['sideA']['player1']['fullName'])
            ws.cell(row, 4, court['sideA']['player2']['fullName'])
            row += 1
            
            # Côté B
            ws.cell(row, 1, court['number'])
            ws.cell(row, 2, 'B')
            ws.cell(row, 3, court['sideB']['player1']['fullName'])
            ws.cell(row, 4, court['sideB']['player2']['fullName'])
            row += 1

# ==================== INIT ====================
def init_db():
    with app.app_context():
        db.create_all()
        
        # Set default settings
        if not Setting.query.filter_by(key='maxTeammates').first():
            Setting.set('maxTeammates', 1)
            Setting.set('maxOpponents', 2)
            Setting.set('maxTeamLevelDiff', 0.4)
            Setting.set('maxMatchLevelDiff', 0.49)

init_db()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)